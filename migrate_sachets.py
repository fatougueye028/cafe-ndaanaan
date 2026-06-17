"""
migrate_sachets.py — Import du stock de sachets depuis Excel
Lit les deux tableaux de l'onglet "4. SUIVI STOCKS" (lignes 27-37).

Usage :
    source venv/bin/activate
    python3 migrate_sachets.py
"""

import openpyxl, gspread, toml
from google.oauth2.service_account import Credentials
from datetime import datetime
import os

EXCEL_PATH   = os.path.join(os.path.dirname(__file__), "..", "Gestion_Cafe_Ndaanaan (1).xlsx")
SECRETS_PATH = os.path.join(os.path.dirname(__file__), ".streamlit", "secrets.toml")

secrets = toml.load(SECRETS_PATH)
scopes  = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
creds   = Credentials.from_service_account_info(secrets["gcp_service_account"], scopes=scopes)
gc      = gspread.authorize(creds)
sh      = gc.open(secrets["SHEET_NAME"])
ws_s    = sh.worksheet("Sachets")

FORMAT_MAP = {"1 kg": "1kg", "500 g": "500g", "250 g": "250g"}
GAMME_MAP  = {"Epicé": "Épicé", "Epice": "Épicé"}
# Mapping couleur → gamme
COULEUR_GAMME = {
    "Blanc":    "Signature",
    "Noir":     "Original",
    "Doré":     "Prestige",
    "Doré vif": "Épicé",
    "Argenté":  "Ñooket",
}

def norm(val, mapping): return mapping.get(str(val).strip(), str(val).strip())
def safe_int(v):
    try: return int(float(v)) if v is not None else 0
    except: return 0
def fmt_date(v):
    if isinstance(v, datetime): return v.strftime("%d/%m/%Y")
    return str(v) if v else ""

wb    = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws_xl = wb["4. SUIVI STOCKS"]

rows = []

# ── Tableau gauche : par gamme (cols 1-4) ──────────────────────
print("=== Sachets par gamme ===")
for r in range(28, 38):
    date_v = ws_xl.cell(r, 1).value
    gamme  = ws_xl.cell(r, 2).value
    fmt    = ws_xl.cell(r, 4).value  # col 4 = Unité (quantité)
    qte    = ws_xl.cell(r, 4).value
    # Corriger : col 3 = Format, col 4 = Unité (stock actuel)
    fmt    = ws_xl.cell(r, 3).value
    qte    = ws_xl.cell(r, 4).value

    if not gamme: continue

    gamme_s = norm(gamme, GAMME_MAP)
    fmt_s   = norm(fmt, FORMAT_MAP)
    stock   = safe_int(qte)

    # Déterminer la couleur selon la gamme
    couleur = {v: k for k, v in COULEUR_GAMME.items()}.get(gamme_s, "")

    row = [
        fmt_date(date_v), couleur, fmt_s, gamme_s,
        stock,  # Qte_Achetee (on utilise le stock connu comme référence)
        0,      # Qte_Utilisee (inconnu)
        stock,  # Stock_Restant
        0,      # Prix_Unitaire (inconnu)
        "Importé depuis Excel — tableau par gamme",
    ]
    rows.append(row)
    print(f"  ✔  {gamme_s} {fmt_s} ({couleur}) → Stock: {stock}")

# ── Tableau droite : par couleur (cols 6-9+) ───────────────────
print("\n=== Sachets par couleur ===")
for r in range(28, 38):
    date_v  = ws_xl.cell(r, 6).value
    couleur = ws_xl.cell(r, 7).value
    fmt     = ws_xl.cell(r, 8).value
    qte_ach = ws_xl.cell(r, 9).value
    restant = ws_xl.cell(r, 11).value  # col 11 = restants

    if not couleur: continue

    couleur_s = str(couleur).strip()
    fmt_s     = norm(fmt, FORMAT_MAP)
    gamme_s   = COULEUR_GAMME.get(couleur_s, "")
    qte_a     = safe_int(qte_ach)
    rest_s    = safe_int(restant)

    row = [
        fmt_date(date_v), couleur_s, fmt_s, gamme_s,
        qte_a,              # Qte_Achetee
        max(0, qte_a - rest_s),  # Qte_Utilisee
        rest_s,             # Stock_Restant
        0,                  # Prix_Unitaire
        "Importé depuis Excel — tableau par couleur",
    ]
    rows.append(row)
    print(f"  ✔  {couleur_s} {fmt_s} ({gamme_s}) → Achetés:{qte_a} Restants:{rest_s}")

# ── Import en une requête ───────────────────────────────────────
if rows:
    ws_s.append_rows(rows, value_input_option="RAW")
    print(f"\n🎉 {len(rows)} ligne(s) sachets importées !")
else:
    print("\nℹ️  Rien à importer.")

print(f"\n📊 https://docs.google.com/spreadsheets/d/{sh.id}/edit")
