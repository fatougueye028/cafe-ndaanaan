"""
migrate_production.py — Import de l'historique de production
Lit l'onglet "1. PRODUCTION" du fichier Excel et l'importe dans Google Sheets.

Usage :
    source venv/bin/activate
    python3 migrate_production.py
"""

import openpyxl
import gspread
import toml
from google.oauth2.service_account import Credentials
from datetime import datetime
import os, sys

# ── Chemins ─────────────────────────────────────────────────────
EXCEL_PATH  = os.path.join(os.path.dirname(__file__), "..", "Gestion_Cafe_Ndaanaan (1).xlsx")
SECRETS_PATH = os.path.join(os.path.dirname(__file__), ".streamlit", "secrets.toml")

# ── Connexion Google Sheets ─────────────────────────────────────
secrets = toml.load(SECRETS_PATH)
scopes  = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
creds   = Credentials.from_service_account_info(secrets["gcp_service_account"], scopes=scopes)
gc      = gspread.authorize(creds)
sh      = gc.open(secrets["SHEET_NAME"])
ws_prod = sh.worksheet("Production")

# ── Lecture Excel ───────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb["1. PRODUCTION"]

# Mapping index de ligne → nom du champ
# (basé sur la structure réelle du fichier)
ROW_MAP = {
    4:  "date",
    5:  "cafe_brut_kg",
    6:  "jar_kg",
    7:  "clous_fcfa",
    8:  "poivre_fcfa",
    9:  "gingembre_fcfa",
    10: "prix_cafe_brut",
    11: "cout_cafe_brut",
    12: "prix_jar",
    13: "cout_jar",
    14: "frais_torref",
    15: "frais_transport",
    16: "sachets_fcfa",
    17: "main_oeuvre",
    18: "affiches_fcfa",
    19: "emballage_fcfa",
    20: "marketing_fcfa",
    21: "cafe_net_kg",
    22: "cout_total",
    23: "cout_revient_kg",
}

# Récupérer les en-têtes (ligne 3 = lots)
header_row = [ws.cell(3, c).value for c in range(1, ws.max_column + 1)]

# Mapping nom de lot → gamme
def lot_to_gamme(lot_name: str) -> str:
    if not lot_name:
        return ""
    n = str(lot_name).upper()
    if n.endswith("S"): return "Signature"
    if n.endswith("O"): return "Original"
    if n.endswith("P"): return "Prestige"
    if n.endswith("E"): return "Épicé"
    if n.endswith("N"): return "Ñooket"
    return "Signature"  # Lot 4 = Lot global, on ignore

def lot_num(lot_name: str) -> str:
    """Lot #2S → Lot 2"""
    if not lot_name:
        return ""
    s = str(lot_name).replace("#", "").strip()
    # Enlever le suffixe de gamme
    if s[-1].upper() in "SOPNE":
        return "Lot " + s[4:-1].strip()
    return s

def safe_float(val) -> float:
    if val is None:
        return 0.0
    try:
        return float(val)
    except:
        return 0.0

def safe_date(val) -> str:
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y")
    if isinstance(val, str) and val != "-":
        return val
    return ""

# ── Construire les lignes ───────────────────────────────────────
rows_to_insert = []

for col_idx, lot_name in enumerate(header_row, start=1):
    if not lot_name or lot_name in ("Composants / Indicateurs", "Unité", "Notes Techniques"):
        continue
    if str(lot_name).startswith("Lot") or str(lot_name).startswith("LOT"):
        # Ignorer Lot 4 (lot global, pas encore produit)
        if "4" in str(lot_name) and not any(c.isalpha() and c.upper() in "SOPNE"
                                             for c in str(lot_name)):
            continue

        data = {}
        for row_idx, field in ROW_MAP.items():
            data[field] = ws.cell(row_idx, col_idx).value

        gamme   = lot_to_gamme(lot_name)
        lot_n   = lot_num(lot_name)
        date_v  = safe_date(data.get("date"))

        cafe_brut_kg    = round(safe_float(data.get("cafe_brut_kg")), 1)
        prix_cafe_brut  = round(safe_float(data.get("prix_cafe_brut")))
        cout_cafe_brut  = round(safe_float(data.get("cout_cafe_brut")))
        jar_kg          = round(safe_float(data.get("jar_kg")), 1)
        prix_jar        = round(safe_float(data.get("prix_jar")))
        cout_jar        = round(safe_float(data.get("cout_jar")))
        clous_fcfa      = round(safe_float(data.get("clous_fcfa")))
        poivre_fcfa     = round(safe_float(data.get("poivre_fcfa")))
        gingembre_fcfa  = round(safe_float(data.get("gingembre_fcfa")))
        frais_torref    = round(safe_float(data.get("frais_torref")))
        frais_transport = round(safe_float(data.get("frais_transport")))
        sachets_fcfa    = round(safe_float(data.get("sachets_fcfa")))
        main_oeuvre     = round(safe_float(data.get("main_oeuvre")))
        affiches_fcfa   = round(safe_float(data.get("affiches_fcfa")))
        emballage_fcfa  = round(safe_float(data.get("emballage_fcfa")))
        marketing_fcfa  = round(safe_float(data.get("marketing_fcfa")))
        cafe_net_kg     = round(safe_float(data.get("cafe_net_kg")), 1)
        cout_total      = round(safe_float(data.get("cout_total")))
        cout_revient_kg = round(safe_float(data.get("cout_revient_kg")), 2)

        row = [
            lot_n, gamme, date_v,
            cafe_brut_kg, prix_cafe_brut, cout_cafe_brut,
            jar_kg, prix_jar, cout_jar,
            clous_fcfa, poivre_fcfa, gingembre_fcfa,
            frais_torref, frais_transport, sachets_fcfa,
            main_oeuvre, affiches_fcfa, emballage_fcfa, marketing_fcfa,
            cafe_net_kg, cout_total, cout_revient_kg,
            0, 0, 0,  # Qte_250g, Qte_500g, Qte_1kg (non disponible dans Excel)
            "",       # Notes
        ]
        rows_to_insert.append(row)
        print(f"  ✔  {lot_name} → {lot_n} / {gamme} | Coût total: {cout_total:,.0f} FCFA | Revient: {cout_revient_kg:,.0f} FCFA/kg")

# ── Vérifier doublons ───────────────────────────────────────────
existing = ws_prod.get_all_records()
existing_keys = {(r.get("Lot",""), r.get("Gamme","")) for r in existing}

new_rows = []
for row in rows_to_insert:
    key = (row[0], row[1])
    if key in existing_keys:
        print(f"  ⏭  Ignoré (déjà présent) : {row[0]} / {row[1]}")
    else:
        new_rows.append(row)

# ── Import ──────────────────────────────────────────────────────
if new_rows:
    ws_prod.append_rows(new_rows, value_input_option="RAW")
    print(f"\n🎉 {len(new_rows)} sous-lot(s) importé(s) avec succès !")
else:
    print("\nℹ️  Aucune nouvelle donnée à importer.")

print(f"\n📊 Google Sheet : https://docs.google.com/spreadsheets/d/{sh.id}/edit")
