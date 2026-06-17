"""
migrate_livreurs.py — Import de l'historique livreurs depuis Excel

Usage :
    source venv/bin/activate
    python3 migrate_livreurs.py
"""

import openpyxl, gspread, toml
from google.oauth2.service_account import Credentials
from datetime import datetime
import os

EXCEL_PATH   = os.path.join(os.path.dirname(__file__), "..", "Gestion_Cafe_Ndaanaan (1).xlsx")
SECRETS_PATH = os.path.join(os.path.dirname(__file__), ".streamlit", "secrets.toml")

secrets = toml.load(SECRETS_PATH)
scopes  = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
creds   = Credentials.from_service_account_info(secrets["gcp_service_account"], scopes=scopes)
gc      = gspread.authorize(creds)
sh      = gc.open(secrets["SHEET_NAME"])

def safe_int(v):
    try: return int(float(v)) if v is not None else 0
    except: return 0
def fmt_date(v):
    if isinstance(v, datetime): return v.strftime("%d/%m/%Y")
    return str(v).strip() if v else ""

wb    = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws_xl = wb["5. LOGISTIQUE LIVREURS"]

# ── Courses Thierno (lignes 4-18) ───────────────────────────────
print("=== Import courses livreurs ===")
ws_liv = sh.worksheet("Livreurs")

existing = ws_liv.get_all_records(value_render_option="UNFORMATTED_VALUE")
existing_keys = {(str(r.get("Livreur","")), str(r.get("Date","")), str(r.get("Zone","")))
                 for r in existing}

rows = []
for r in range(4, 25):
    livreur = ws_xl.cell(r, 1).value
    date_v  = ws_xl.cell(r, 2).value
    zone    = ws_xl.cell(r, 3).value
    tarif   = ws_xl.cell(r, 4).value
    tel     = ws_xl.cell(r, 5).value
    courses = ws_xl.cell(r, 6).value
    montant = ws_xl.cell(r, 7).value

    if not livreur or not date_v or str(livreur).startswith("TOTAL"):
        continue

    livreur_s = str(livreur).strip()
    date_s    = fmt_date(date_v)
    zone_s    = str(zone).strip() if zone else ""
    tarif_s   = safe_int(tarif)
    nb        = safe_int(courses) if courses else 1
    montant_s = safe_int(montant)

    key = (livreur_s, date_s, zone_s)
    if key in existing_keys:
        print(f"  ⏭  Ignoré (existe): {livreur_s} {date_s} {zone_s}")
        continue

    rows.append([livreur_s, date_s, zone_s, tarif_s, nb, montant_s])
    print(f"  ✔  {livreur_s} | {date_s} | {zone_s} → {montant_s} FCFA")

if rows:
    ws_liv.append_rows(rows, value_input_option="RAW")
    print(f"\n✅ {len(rows)} course(s) importée(s).")
else:
    print("\nℹ️  Déjà à jour.")

print(f"📊 Total : {sum(r[5] for r in rows):,} FCFA importés")

# ── Contacts Transporteurs ──────────────────────────────────────
print("\n=== Import contacts transporteurs ===")
ws_trans = sh.worksheet("Transporteurs")

existing_trans = {r.get("Nom","") for r in ws_trans.get_all_records()}

transporteurs = [
    ["GP Canada",       "Livraison locale", "Ouest Foire",                    "",                         "78 170 51 51 / 77 817 50 50 / 77 427 18 18", "",              ""],
    ["Fred France",     "Export France",    "Guédiawaye → France",            "5 euros/kg (~54.5€ total)", "78 896 12 93",                               "",              ""],
    ["Bateau France",   "Export France",    "Dakar → France (maritime)",       "3 euros/kg",               "",                                           "1 mois environ", ""],
    ["Chrono Poste",    "Export France",    "Dakar → France",                  "",                         "",                                           "",              ""],
    ["Mondial Relay",   "Export France",    "France (point relais)",           "",                         "",                                           "",              ""],
]

new_trans = [t for t in transporteurs if t[0] not in existing_trans]
if new_trans:
    ws_trans.append_rows(new_trans, value_input_option="RAW")
    print(f"  ✅ {len(new_trans)} transporteur(s) importé(s).")
else:
    print("  ℹ️  Déjà à jour.")

print(f"\n🎉 Migration livreurs terminée !")
print(f"📊 https://docs.google.com/spreadsheets/d/{sh.id}/edit")
