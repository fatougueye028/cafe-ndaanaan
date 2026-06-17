"""
migrate_stock.py — Import de l'historique de stock depuis Excel
Lit "4. SUIVI STOCKS" et importe dans Google Sheets.

Usage :
    source venv/bin/activate
    python3 migrate_stock.py
"""

import openpyxl
import gspread
import toml
from google.oauth2.service_account import Credentials
from datetime import datetime
import os, sys

EXCEL_PATH   = os.path.join(os.path.dirname(__file__), "..", "Gestion_Cafe_Ndaanaan (1).xlsx")
SECRETS_PATH = os.path.join(os.path.dirname(__file__), ".streamlit", "secrets.toml")

secrets = toml.load(SECRETS_PATH)
scopes  = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
creds   = Credentials.from_service_account_info(secrets["gcp_service_account"], scopes=scopes)
gc      = gspread.authorize(creds)
sh      = gc.open(secrets["SHEET_NAME"])

# ── Normalisation ───────────────────────────────────────────────
FORMAT_MAP = {"1 kg": "1kg", "500 g": "500g", "250 g": "250g",
              "1kg": "1kg", "500g": "500g", "250g": "250g"}
GAMME_MAP  = {"Epicé": "Épicé", "Epice": "Épicé", "Nooket": "Ñooket"}

def norm_format(v):
    return FORMAT_MAP.get(str(v).strip(), str(v).strip())

def norm_gamme(v):
    s = str(v).strip()
    return GAMME_MAP.get(s, s)

def safe_int(v):
    try: return int(float(v)) if v is not None else 0
    except: return 0

def fmt_date(v):
    if isinstance(v, datetime): return v.strftime("%d/%m/%Y")
    return str(v) if v else ""

# ── Lire l'Excel ────────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb["4. SUIVI STOCKS"]

# ── PARTIE 1 : Historique par lot (lignes 4-23) ─────────────────
print("\n=== Import historique lots ===")
ws_hist = sh.worksheet("Stock_Historique")
existing_hist = {(r["Lot"], r["Gamme"], r["Format"]) for r in ws_hist.get_all_records()}

rows_hist = []
for r in range(4, 25):
    date_v = ws.cell(r, 1).value
    lot    = ws.cell(r, 2).value
    gamme  = ws.cell(r, 3).value
    fmt    = ws.cell(r, 4).value
    prod   = ws.cell(r, 5).value
    cmd    = ws.cell(r, 6).value
    vend   = ws.cell(r, 7).value
    rest   = ws.cell(r, 8).value
    comm   = ws.cell(r, 9).value

    if not lot or not gamme: continue

    lot_s   = str(lot).strip()
    gamme_s = norm_gamme(gamme)
    fmt_s   = norm_format(fmt)
    key     = (lot_s, gamme_s, fmt_s)

    if key in existing_hist:
        print(f"  ⏭  Ignoré (existe): {lot_s} / {gamme_s} / {fmt_s}")
        continue

    rows_hist.append([
        fmt_date(date_v), lot_s, gamme_s, fmt_s,
        safe_int(prod), safe_int(cmd), safe_int(vend), safe_int(rest),
        str(comm) if comm else ""
    ])
    print(f"  ✔  {lot_s} / {gamme_s} / {fmt_s} → Prod:{safe_int(prod)} Cmd:{safe_int(cmd)} Vend:{safe_int(vend)} Rest:{safe_int(rest)}")

if rows_hist:
    ws_hist.append_rows(rows_hist, value_input_option="RAW")
    print(f"\n✅ {len(rows_hist)} ligne(s) historique importées.")
else:
    print("\nℹ️  Historique déjà à jour.")

# ── PARTIE 2 : Mettre à jour le stock actuel ────────────────────
print("\n=== Mise à jour stock actuel ===")
ws_stock = sh.worksheet("Stock")

# Agréger le stock restant par Gamme+Format depuis l'historique
from collections import defaultdict
stock_agg = defaultdict(lambda: {"prod": 0, "cmd": 0, "vend": 0, "rest": 0})

for r in range(4, 25):
    gamme = ws.cell(r, 3).value
    fmt   = ws.cell(r, 4).value
    prod  = ws.cell(r, 5).value
    cmd   = ws.cell(r, 6).value
    vend  = ws.cell(r, 7).value
    rest  = ws.cell(r, 8).value
    if not gamme: continue
    key = (norm_gamme(gamme), norm_format(fmt))
    stock_agg[key]["prod"] += safe_int(prod)
    stock_agg[key]["cmd"]  += safe_int(cmd)
    stock_agg[key]["vend"] += safe_int(vend)
    stock_agg[key]["rest"] += safe_int(rest)

# Lire le stock actuel
df_stock = ws_stock.get_all_records(value_render_option="UNFORMATTED_VALUE")
headers  = ws_stock.row_values(1)

col_prod = headers.index("Unites_Produites") + 1 if "Unites_Produites" in headers else None
col_cmd  = headers.index("Unites_Commandees") + 1 if "Unites_Commandees" in headers else None
col_vend = headers.index("Unites_Vendues") + 1 if "Unites_Vendues" in headers else None
col_rest = headers.index("Stock_Restant") + 1 if "Stock_Restant" in headers else None
col_maj  = headers.index("Derniere_MAJ") + 1 if "Derniere_MAJ" in headers else None

today = datetime.today().strftime("%d/%m/%Y")

updated = 0
for i, row in enumerate(df_stock):
    gamme = str(row.get("Gamme", "")).strip()
    fmt   = str(row.get("Format", "")).strip()
    loc   = str(row.get("Localisation", "")).strip()

    if loc not in ("Touba", "Dakar"):  # Stock historique = Touba/Dakar
        continue

    key = (gamme, fmt)
    if key in stock_agg:
        data = stock_agg[key]
        sheet_row = i + 2
        if col_prod: ws_stock.update_cell(sheet_row, col_prod, data["prod"])
        if col_cmd:  ws_stock.update_cell(sheet_row, col_cmd,  data["cmd"])
        if col_vend: ws_stock.update_cell(sheet_row, col_vend, data["vend"])
        if col_rest: ws_stock.update_cell(sheet_row, col_rest, data["rest"])
        if col_maj:  ws_stock.update_cell(sheet_row, col_maj,  today)
        print(f"  ✔  {gamme} {fmt} ({loc}) → Prod:{data['prod']} Cmd:{data['cmd']} Vend:{data['vend']} Rest:{data['rest']}")
        updated += 1

print(f"\n✅ {updated} ligne(s) stock mises à jour.")
print(f"\n📊 Google Sheet : https://docs.google.com/spreadsheets/d/{sh.id}/edit")
