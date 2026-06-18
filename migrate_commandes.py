"""
migrate_commandes.py — Import de l'historique des commandes depuis Excel

Sources :
  - Gestion_Cafe_Ndaanaan.xlsx → "2. CARNET COMMANDES" (212 commandes Sénégal)
  - Kafe Ndaanaan France.xlsx  → "Feuille 1" (39 commandes France)

Usage :
    source venv/bin/activate
    python3 migrate_commandes.py
"""

import openpyxl, gspread, toml
from google.oauth2.service_account import Credentials
from datetime import datetime
import os

EXCEL_SN  = os.path.join(os.path.dirname(__file__), "..", "Gestion_Cafe_Ndaanaan (1).xlsx")
EXCEL_FR  = os.path.join(os.path.dirname(__file__), "..", "Kafe Ndaanaan France.xlsx")
SECRETS   = os.path.join(os.path.dirname(__file__), ".streamlit", "secrets.toml")

secrets = toml.load(SECRETS)
scopes  = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
creds   = Credentials.from_service_account_info(secrets["gcp_service_account"], scopes=scopes)
gc      = gspread.authorize(creds)
sh      = gc.open(secrets["SHEET_NAME"])
ws      = sh.worksheet("Commandes")

# ── Normalisation ───────────────────────────────────────────────
FORMAT_MAP  = {"500 g": "500g", "1 kg": "1kg", "250 g": "250g",
               "500g": "500g", "1kg": "1kg", "250g": "250g"}
GAMME_MAP   = {"Epicé": "Épicé", "Epice": "Épicé", "Nooket": "Ñooket",
               "Épicé": "Épicé", "Ñooket": "Ñooket"}
LIV_MAP     = {"Livré": "Livrée", "Livree": "Livrée", "livrée": "Livrée",
               "En attente": "Commande confirmée", "Planifié": "Commande confirmée",
               "À préparer": "Commande confirmée", "Non livré": "Commande confirmée",
               "Annulé": "Annulée", "Annulée": "Annulée", "": "Commande confirmée"}
# Mapping Type_Demande → Statut_Livraison
STATUT_LIV_MAP = {
    "Livrée":             "Livrée",
    "Préparée":           "Préparée",
    "Commande confirmée": "À préparer",
    "Annulée":            "Annulée",
}
PAY_MAP     = {"Payé": "Payé", "Paye": "Payé", "payé": "Payé",
               "En attente": "Non payé", "Non payé": "Non payé",
               "Partiel": "Partiel", "": "Non payé"}

def norm_format(v):  return FORMAT_MAP.get(str(v).strip(), str(v).strip())
def norm_gamme(v):   return GAMME_MAP.get(str(v).strip(), str(v).strip())
def norm_liv(v):     return LIV_MAP.get(str(v).strip(), "À préparer")
def norm_pay(v):     return PAY_MAP.get(str(v).strip(), "Non payé")
def safe_float(v):
    try: return float(v) if v is not None else 0.0
    except: return 0.0
def safe_int(v):
    try: return int(float(v)) if v is not None else 0
    except: return 0
def fmt_date(v):
    if isinstance(v, datetime): return v.strftime("%d/%m/%Y")
    return str(v).strip() if v else ""

# Lire IDs existants
existing = ws.get_all_records(value_render_option="UNFORMATTED_VALUE")
# Clé composite : ID + Gamme + Format (une commande peut avoir plusieurs produits)
existing_keys = {(str(r.get("ID","")), str(r.get("Gamme","")), str(r.get("Format","")))
                 for r in existing}
print(f"Lignes déjà présentes : {len(existing_keys)}")

rows = []

# ── SÉNÉGAL (212 commandes) ─────────────────────────────────────
print("\n=== Import commandes Sénégal ===")
wb_sn = openpyxl.load_workbook(EXCEL_SN, data_only=True)
ws_sn = wb_sn["2. CARNET COMMANDES"]

for r in range(4, ws_sn.max_row + 1):
    date_v  = ws_sn.cell(r, 1).value
    cmd_id  = ws_sn.cell(r, 2).value
    client  = ws_sn.cell(r, 3).value
    tel     = ws_sn.cell(r, 4).value
    lot     = ws_sn.cell(r, 5).value
    gamme   = ws_sn.cell(r, 6).value
    fmt     = ws_sn.cell(r, 7).value
    qte     = ws_sn.cell(r, 8).value
    prix    = ws_sn.cell(r, 9).value
    ca      = ws_sn.cell(r, 10).value
    liv     = ws_sn.cell(r, 11).value
    pay     = ws_sn.cell(r, 12).value
    comm    = ws_sn.cell(r, 13).value

    if not cmd_id or not str(cmd_id).startswith("CMD"):
        continue
    gamme_n = norm_gamme(gamme or "")
    fmt_n   = norm_format(fmt or "")
    if (str(cmd_id).strip(), gamme_n, fmt_n) in existing_keys:
        continue

    # Déterminer la zone
    lot_s = str(lot).strip() if lot else ""
    zone  = "Touba" if "Touba" in lot_s or "touba" in lot_s.lower() else "Dakar"

    rows.append([
        fmt_date(date_v),
        str(cmd_id).strip(),
        str(client).strip() if client else "",
        str(tel).strip() if tel else "",
        "",                          # Adresse
        zone,
        gamme_n,
        fmt_n,
        safe_int(qte),
        safe_float(prix),
        safe_float(ca),
        "FCFA",
        norm_liv(liv or ""),                              # Type_Demande
        STATUT_LIV_MAP.get(norm_liv(liv or ""), "À préparer"),  # Statut_Livraison
        norm_pay(pay or ""),
        "",                          # Source
        lot_s,
        str(comm).strip() if comm else "",
        "",                          # Date_Prevue
    ])

print(f"  ✔  {len(rows)} lignes Sénégal à importer")

# ── Import en batch ─────────────────────────────────────────────
if rows:
    # Envoyer par batch de 50 pour éviter timeout
    batch_size = 50
    total = 0
    for i in range(0, len(rows), batch_size):
        batch = rows[i:i + batch_size]
        ws.append_rows(batch, value_input_option="RAW")
        total += len(batch)
        print(f"  📤 {total}/{len(rows)} lignes envoyées...")
    print(f"\n🎉 {len(rows)} commandes importées !")
else:
    print("\nℹ️  Toutes les commandes sont déjà présentes.")

print(f"\n📊 https://docs.google.com/spreadsheets/d/{sh.id}/edit")
